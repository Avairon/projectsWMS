from flask import Blueprint, render_template, jsonify, send_file, request
import io
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask_login import login_required, current_user
from app.utils import load_data
from app.tables import create_projects_table, create_tasks_table
from config import Config

app_config = Config()
reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/reports')
@login_required
def reports():
    """Страница отчетов для администраторов"""
    if current_user.role != 'admin':
        from flask import flash, redirect, url_for
        flash('У вас нет прав доступа к этой странице')
        return redirect(url_for('dashboard.dashboard'))
    return render_template('reports.html')


@reports_bp.route('/api/reports/projects')
@login_required
def get_projects_report():
    """
    API endpoint для получения данных отчета по проектам
    
    Параметры запроса:
        search: строка для поиска по текстовым полям
        start_date: начальная дата для фильтра (формат DD.MM.YYYY)
        end_date: конечная дата для фильтра (формат DD.MM.YYYY)
        date_field: поле с датой для фильтра ('start_date' или 'end_date')
    """
    if current_user.role != 'admin':
        return jsonify({'error': 'У вас нет прав доступа к этой странице'}), 403
    
    projects = load_data(app_config.PROJECTS_DB)
    users = load_data(app_config.USERS_DB)
    
    # Создаем таблицу
    table = create_projects_table(projects, users)
    
    # Применяем поиск, если указан
    search_query = request.args.get('search', '')
    if search_query:
        table.search(search_query)
    
    # Применяем фильтр по диапазону дат, если указаны
    date_field = request.args.get('date_field')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if date_field and start_date and end_date:
        table.filter_by_date_range(date_field, start_date, end_date)
    elif date_field and start_date:
        # Фильтр по одной дате
        table.filter_by_date(date_field, start_date)
    
    # Возвращаем данные в формате для фронтенда
    return jsonify(table.to_dict())


@reports_bp.route('/api/reports/tasks')
@login_required
def get_tasks_report():
    """
    API endpoint для получения данных отчета по задачам проектов
    
    Параметры запроса:
        search: строка для поиска по текстовым полям
        start_date: начальная дата для фильтра (формат DD.MM.YYYY)
        end_date: конечная дата для фильтра (формат DD.MM.YYYY)
        date_field: поле с датой для фильтра ('start_date' или 'end_date')
    """
    if current_user.role != 'admin':
        return jsonify({'error': 'У вас нет прав доступа к этой странице'}), 403
    
    tasks = load_data(app_config.TASKS_DB)
    users = load_data(app_config.USERS_DB)
    projects = load_data(app_config.PROJECTS_DB)
    
    # Создаем таблицу
    table = create_tasks_table(tasks, users, projects)
    
    # Применяем поиск, если указан
    search_query = request.args.get('search', '')
    if search_query:
        table.search(search_query)
    
    # Применяем фильтр по диапазону дат, если указаны
    date_field = request.args.get('date_field')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if date_field and start_date and end_date:
        table.filter_by_date_range(date_field, start_date, end_date)
    elif date_field and start_date:
        # Фильтр по одной дате
        table.filter_by_date(date_field, start_date)
    
    # Возвращаем данные в формате для фронтенда
    return jsonify(table.to_dict())


@reports_bp.route('/api/reports/projects/download')
@login_required
def download_projects_report():
    """Download endpoint for projects report in Excel format"""
    if current_user.role != 'admin':
        return jsonify({'error': 'У вас нет прав доступа к этой странице'}), 403
    
    projects = load_data(app_config.PROJECTS_DB)
    users = load_data(app_config.USERS_DB)
    
    # Получаем параметры фильтрации из запроса
    search_query = request.args.get('search', '')
    date_field = request.args.get('date_field')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Создаем таблицу и применяем фильтры
    table = create_projects_table(projects, users)
    
    if search_query:
        table.search(search_query)
    
    if date_field and start_date and end_date:
        table.filter_by_date_range(date_field, start_date, end_date)
    elif date_field and start_date:
        table.filter_by_date(date_field, start_date)
    
    # Получаем отфильтрованные данные
    filtered_data = table.get_filtered_data()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по проектам"
    
    # Define headers (с нумерацией)
    headers = ["#", "Дата начала", "Дата окончания", "Название проекта", "Статус", "Куратор проекта", "Руководитель проекта"]
    
    # Заголовки таблицы
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Данные таблицы
    for row_num, item in enumerate(filtered_data, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)  # Номер строки
        ws.cell(row=row_num, column=2, value=item.get('start_date', ''))
        ws.cell(row=row_num, column=3, value=item.get('end_date', ''))
        ws.cell(row=row_num, column=4, value=item.get('project_name', ''))
        ws.cell(row=row_num, column=5, value=item.get('status', ''))
        ws.cell(row=row_num, column=6, value=item.get('curator_name', ''))
        ws.cell(row=row_num, column=7, value=item.get('manager_name', ''))
    
    # Стилизация строк
    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2, max_row=len(filtered_data) + 1):
        fill = even_fill if row[0].row % 2 == 0 else odd_fill
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Информация о фильтрах внизу
    if search_query or date_field:
        info_row = len(filtered_data) + 3
        ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=7)
        info_cell = ws.cell(row=info_row, column=1)
        info_text = f"Всего записей: {table.get_total_count()}, Отображено: {table.get_filtered_count()}"
        if search_query:
            info_text += f" | Поиск: '{search_query}'"
        if date_field and start_date and end_date:
            info_text += f" | Фильтр: {date_field} от {start_date} до {end_date}"
        elif date_field and start_date:
            info_text += f" | Фильтр: {date_field} = {start_date}"
        info_cell.value = info_text
        info_cell.font = Font(italic=True, size=9)
        info_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        info_cell.alignment = Alignment(horizontal="left")
    
    # Save to BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = int(time.time())
    filename = f'report_projects_{current_user.id}_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/api/reports/tasks/download')
@login_required
def download_tasks_report():
    """Download endpoint for tasks report in Excel format"""
    if current_user.role != 'admin':
        return jsonify({'error': 'У вас нет прав доступа к этой странице'}), 403
    
    tasks = load_data(app_config.TASKS_DB)
    users = load_data(app_config.USERS_DB)
    projects = load_data(app_config.PROJECTS_DB)
    
    # Получаем параметры фильтрации из запроса
    search_query = request.args.get('search', '')
    date_field = request.args.get('date_field')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Создаем таблицу и применяем фильтры
    table = create_tasks_table(tasks, users, projects)
    
    if search_query:
        table.search(search_query)
    
    if date_field and start_date and end_date:
        table.filter_by_date_range(date_field, start_date, end_date)
    elif date_field and start_date:
        table.filter_by_date(date_field, start_date)
    
    # Получаем отфильтрованные данные
    filtered_data = table.get_filtered_data()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по задачам"
    
    # Define headers (с нумерацией)
    headers = ["#", "Исполнитель", "Проект", "Задача", "Дата начала задачи", "Дата окончания задачи", "Статус задачи"]
    
    # Заголовки таблицы
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Данные таблицы
    for row_num, item in enumerate(filtered_data, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)  # Номер строки
        ws.cell(row=row_num, column=2, value=item.get('executor', ''))
        ws.cell(row=row_num, column=3, value=item.get('project', ''))
        ws.cell(row=row_num, column=4, value=item.get('task', ''))
        ws.cell(row=row_num, column=5, value=item.get('start_date', ''))
        ws.cell(row=row_num, column=6, value=item.get('end_date', ''))
        ws.cell(row=row_num, column=7, value=item.get('status', ''))
    
    # Стилизация строк
    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2, max_row=len(filtered_data) + 1):
        fill = even_fill if row[0].row % 2 == 0 else odd_fill
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Информация о фильтрах внизу
    if search_query or date_field:
        info_row = len(filtered_data) + 3
        ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=7)
        info_cell = ws.cell(row=info_row, column=1)
        info_text = f"Всего записей: {table.get_total_count()}, Отображено: {table.get_filtered_count()}"
        if search_query:
            info_text += f" | Поиск: '{search_query}'"
        if date_field and start_date and end_date:
            info_text += f" | Фильтр: {date_field} от {start_date} до {end_date}"
        elif date_field and start_date:
            info_text += f" | Фильтр: {date_field} = {start_date}"
        info_cell.value = info_text
        info_cell.font = Font(italic=True, size=9)
        info_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        info_cell.alignment = Alignment(horizontal="left")
    
    # Save to BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = int(time.time())
    filename = f'report_tasks_{current_user.id}_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )